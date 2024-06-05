import kivy

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.core.window import Window
from openpyxl import load_workbook

class OraculoApp(App):
    def build(self):
        self.wb = load_workbook("Oraculo_1.0.xlsx")
        self.sheet = self.wb.active

        self.inputs = []

        # Layout principal
        main_layout = BoxLayout(orientation='vertical', padding=50, spacing=2)

        # Layout para os inputs
        input_layout = BoxLayout(orientation='vertical', padding=1, spacing=20)

        # Layout para os botões
        button_layout = BoxLayout(orientation='horizontal', padding=15, spacing=1)

        self.labels = [self.sheet.cell(row=1, column=i).value for i in range(2, 10)]

        for i in range(8):
            # Caixa de input com label, campo de entrada e botões
            input_box = BoxLayout(orientation='horizontal', spacing=1, size_hint=(1, 2), height=50)

            label = Label(text=self.labels[i], size_hint=(1, None), height=10)
            input_field = TextInput(size_hint=(1, 1), height=20)
            button1 = Button(text='1º', size_hint=(1, 1), height=20,
                             on_press=self.create_set_input_value(input_field, 'I'))
            button2 = Button(text='2º', size_hint=(1, 1), height=20,
                             on_press=self.create_set_input_value(input_field, 'II'))

            input_box.add_widget(label)
            input_box.add_widget(input_field)
            input_box.add_widget(button1)
            input_box.add_widget(button2)

            self.inputs.append(input_field)

            if i > 4:
                input_layout.add_widget(input_box)
            else:
                input_layout.add_widget(input_box)

        # Botão de pesquisar
        search_button = Button(text='Pesquisar', size_hint=(0.5, None), height=50, on_press=self.search)

        # Botão de limpar
        clear_button = Button(text='Limpar', size_hint=(0.5, None), height=50, on_press=self.clear_inputs)

        # Adiciona widgets ao layout de botões
        button_layout.add_widget(search_button)
        button_layout.add_widget(clear_button)

        # Label para exibir o resultado da pesquisa
        self.result_label = Label(text='', size_hint=(1, None), height=20)

        # Adiciona widgets ao layout principal
        main_layout.add_widget(input_layout)
        main_layout.add_widget(button_layout)
        main_layout.add_widget(self.result_label)

        return main_layout

    def create_set_input_value(self, input_field, value):
        def set_input_value(instance):
            input_field.text = value
        return set_input_value

    def search(self, instance):
        inputs = [inp.text for inp in self.inputs]
        found = False

        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if list(row[1:9]) == inputs:
                self.result_label.text = row[0]
                found = True
                break

        if not found:
            self.result_label.text = 'Nenhum resultado encontrado'

    def clear_inputs(self, instance):
        for input_field in self.inputs:
            input_field.text = ''

if __name__ == '__main__':
    OraculoApp().run()